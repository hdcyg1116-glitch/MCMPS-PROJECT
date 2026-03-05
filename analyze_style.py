import openpyxl
import json
import os
from openpyxl.utils import get_column_letter

file_path = r"c:\Users\예광\Desktop\새 폴더\MCMPS-PROJECT\(고객송부)남산TC 26년 2월 출하 계획(20260213).xlsx"

def analyze_excel_style(file_path):
    print(f"Analyzing {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 첫 번째 시트 기준 (또는 '생산계획' 등 특정 시트)
    ws = wb.active
    print(f"Active Sheet: {ws.title}")
    
    analysis = {
        'sheet_name': ws.title,
        'dimensions': {
            'max_row': ws.max_row,
            'max_col': ws.max_column
        },
        'columns': {},
        'cells': {},
        'merged_cells': [str(r) for r in ws.merged_cells.ranges]
    }
    
    # 1. 컬럼 너비 분석 (최대 30열까지만 샘플링)
    for col_idx in range(1, min(ws.max_column + 1, 30)):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        analysis['columns'][col_letter] = width
        
    # 2. 상위 20행 x 20열 정도의 주요 서식 샘플링 (헤더와 데이터 샘플)
    for row in range(1, min(ws.max_row + 1, 50)):
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row, column=col)
            
            # 값이 있거나 배경색이 있는 셀만 분석하여 용량 줄이기
            has_value = cell.value is not None
            has_bg = cell.fill.start_color.index != '00000000' if cell.fill else False
            
            if has_value or has_bg:
                font = cell.font
                fill = cell.fill
                border = cell.border
                alignment = cell.alignment
                
                cell_id = f"{get_column_letter(col)}{row}"
                
                # 색상값이 Theme Color인 경우 처리
                fg_color = fill.start_color.rgb if fill and hasattr(fill.start_color, 'rgb') and type(fill.start_color.rgb) == str else None
                font_color = font.color.rgb if font and hasattr(font, 'color') and font.color and hasattr(font.color, 'rgb') and type(font.color.rgb) == str else None
                
                analysis['cells'][cell_id] = {
                    'value': str(cell.value)[:50] if cell.value else None, # 너무 긴 텍스트는 자름
                    'font': {
                        'name': font.name if font else None,
                        'size': font.size if font else None,
                        'bold': font.bold if font else False,
                        'color': font_color
                    },
                    'fill': {
                        'fgColor': fg_color
                    },
                    'alignment': {
                        'horizontal': alignment.horizontal if alignment else None,
                        'vertical': alignment.vertical if alignment else None
                    }
                }

    # 결과를 json으로 저장
    output_path = os.path.join(os.path.dirname(file_path), "excel_style_analysis.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
        
    print(f"Analysis saved to {output_path}")

if __name__ == "__main__":
    analyze_excel_style(file_path)
