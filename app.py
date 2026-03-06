from flask import Flask, render_template, request, jsonify, redirect, url_for
import pandas as pd
import os
import json

app = Flask(__name__)

# 파일 경로 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PLAN = os.path.join(BASE_DIR, '(반출승인)생산2팀_생산진도표(TC)_260223.xlsx')
EXCEL_FILE_DAILY = os.path.join(BASE_DIR, 'TC 대조립 일일 진도현황(260304).xlsx')

# 전역 변수 (캐싱용)
_SHEET_CACHE = {"data": [], "last_mtime": 0}
import threading
import time

def update_cache_bg():
    global _SHEET_CACHE
    while True:
        try:
            if not os.path.exists(EXCEL_FILE_DAILY):
                time.sleep(10)
                continue
                
            mtime = os.path.getmtime(EXCEL_FILE_DAILY)
            if _SHEET_CACHE["last_mtime"] == mtime and _SHEET_CACHE["data"]:
                time.sleep(10)
                continue

            from openpyxl import load_workbook
            wb = load_workbook(EXCEL_FILE_DAILY, read_only=True)
            
            valid_sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.sheet_state == 'visible' and (sheet_name.startswith('TC 호기별') or sheet_name == 'TC 대조립'):
                    valid_sheets.append(sheet_name)
            wb.close()
            
            formatted_sheets = []
            
            if 'TC 대조립' in valid_sheets:
                formatted_sheets.append({
                    "type": "daily",
                    "sheet": "TC 대조립",
                    "display": "TC 대조립"
                })
                valid_sheets.remove('TC 대조립')

            daily_sheets = []
            import re
            for s in valid_sheets:
                match = re.search(r'(\d+)년\s*(\d+)월', s)
                sort_key = ""
                if match:
                    year, month = match.groups()
                    full_year = f"20{year}" if len(year) == 2 else year
                    sort_key = f"{full_year}{int(month):02d}"
                
                daily_sheets.append({
                    "type": "daily",
                    "sheet": s,
                    "display": s,
                    "sort_key": sort_key
                })
            
            daily_sheets.sort(key=lambda x: x['sort_key'], reverse=True)
            formatted_sheets.extend([{k: v for k, v in d.items() if k != 'sort_key'} for d in daily_sheets])
            
            _SHEET_CACHE["data"] = formatted_sheets
            _SHEET_CACHE["last_mtime"] = mtime
            print("Background cache updated successfully.")

        except Exception as e:
            print(f"Error in bg cache worker: {e}")
        
        # 10초마다 엑셀 파일 변경 여부 확인
        time.sleep(10)

# 백그라운드 스레드 시작
threading.Thread(target=update_cache_bg, daemon=True).start()

def get_excel_sheets():
    # 백그라운드 스레드에 의해 즉각적으로 준비된 캐시 반환
    return _SHEET_CACHE["data"] if _SHEET_CACHE["data"] else [
        { "type": "plan", "sheet": None, "display": "출하 계획표 (기본)" }
    ]

def parse_production_data(file_path, sheet_name=0):
    try:
        print(f"Parsing File: {file_path}, Request Sheet: {sheet_name}")
        
        # 0. 가장 적합한 시트와 헤더를 찾는 로직
        header_keywords = ['NO.', 'NO', '생산월', '기종', '호기', '전산기종', '출하처', '목적지', '오더', '진행상태', '현공정', '생산직', '최초출하일', '개정출하일', 'SERIAL']
        
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
        
        best_sheet = sheet_name
        best_header_idx = -1
        max_score = 0
        best_df = None
        
        # 만약 명시적으로 문자열을 전달받지 않았다면(기본값 0 등) 모든 시트를 검토
        sheets_to_scan = [sheet_name] if isinstance(sheet_name, str) else sheet_names
        
        for s in sheets_to_scan:
            try:
                # xl 객체를 사용하여 파일 I/O 비용 최소화
                temp_df = xl.parse(sheet_name=s, header=None)
                scan_limit = min(30, len(temp_df))
                for idx in range(scan_limit):
                    row = temp_df.iloc[idx]
                    row_str = " ".join([str(val) for val in row if pd.notna(val)]).replace(' ', '').upper()
                    
                    score = 0
                    for kw in header_keywords:
                        if kw.replace(' ', '').upper() in row_str:
                            score += 1
                    
                    if score > max_score:
                        max_score = score
                        best_sheet = s
                        best_header_idx = idx
                        best_df = temp_df
            except Exception as e:
                print(f"Sheet {s} scan error: {e}")
                continue

        # 만약 적합한 시트를 찾지 못했다면 원래 요청된 시트를 사용
        if best_df is None:
            df = xl.parse(sheet_name=sheet_name, header=1)
            # 폴백
            is_monthly_sheet = str(sheet_name).startswith('TC 호기별')
            if not is_monthly_sheet:
                if len(df.columns) < 5 or ("Unnamed" in str(df.columns[0]) and "Unnamed" in str(df.columns[1])):
                    df = xl.parse(sheet_name=sheet_name, header=0)
            best_df = df
            
            # 여기서부터는 기존 헤더 탐색 로직
            best_header_idx = -1
            best_score = 0
            scan_limit = min(30, len(best_df))
            for idx in range(scan_limit):
                row = best_df.iloc[idx]
                row_str = " ".join([str(val) for val in row if pd.notna(val)]).replace(' ', '').upper()
                score = sum(1 for kw in header_keywords if kw.replace(' ', '').upper() in row_str)
                if score > best_score and score >= 3:
                    best_score = score
                    best_header_idx = idx

        df = best_df
        print(f"Selected Sheet: {best_sheet}, Identified Header Row: {best_header_idx}")

        col_count = len(df.columns)
        row_count = len(df)
        print(f"Excel Structure: {row_count} rows, {col_count} columns")

        if col_count < 5:
            print(f"Error: Too few columns ({col_count}).")
            return []

        # 동적 헤더 탐색 알고리즘
        col_map = {}
        header_idx = best_header_idx
        
        if header_idx != -1:
            header_row = df.iloc[header_idx]
            for col_idx, val in enumerate(header_row):
                if pd.notna(val):
                    clean_name = str(val).replace('\n', '').replace(' ', '').upper()
                    if clean_name:
                        col_map[clean_name] = col_idx
        else:
            # 적절한 헤더를 찾지 못했다면 기존 df.columns가 헤더인지 확인
            columns_str = " ".join([str(c) for c in df.columns]).upper()
            if 'NO.' in columns_str or 'NO ' in columns_str or '생산월' in columns_str:
                for col_idx, val in enumerate(df.columns):
                    clean_name = str(val).replace('\n', '').replace(' ', '').upper()
                    col_map[clean_name] = col_idx

        def get_idx(*possible_names, default=None):
            for name in possible_names:
                clean_target = name.upper().replace(' ', '')
                if clean_target in col_map:
                    return col_map[clean_target]
                # 컬럼명에 포함되어 있는지 부분 일치 검색
                for map_key in col_map.keys():
                    if clean_target in map_key:
                        return col_map[map_key]
            return default

        # 검색어 동의어 사전 대폭 확대
        no_idx = get_idx('NO.', 'NO', '번호', '순번', default=1)
        month_idx = get_idx('생산월', '월', 'MONTH', default=2)
        section_idx = get_idx('생산직', '직', '반', 'SECTION', default=3)
        model_idx = get_idx('기종', '전산기종', '모델', 'MODEL', '품명', default=5)
        serial_idx_mapped = get_idx('호기', '시리얼', 'SERIAL', 'S/N', default=-1)
        order_idx = get_idx('오더', 'ORDER', '작업지시', '작업오더', default=7)
        customer_idx = get_idx('출하처', '목적지', '고객', 'CUSTOMER', '납품처', default=8)
        first_shipment_idx = get_idx('최초출하일', '최조출하일', '초기출하', default=9)
        target_idx = get_idx('개정출하일', '출하일', '목표일', '수정출하일', default=10)
        base_idx = get_idx('BASE시작일', 'BASE작업일', '베이스', 'BASE', default=11)
        first_start_idx = get_idx('최초시작일', '시작일', '초기시작', default=12)
        revised_start_idx = get_idx('개정시작일', '수정시작', default=13)
        nc_idx = get_idx('NC', '엔씨', default=14)
        status_idx = get_idx('현공정', '진행상태', '상태', '공정', 'STATUS', default=15)
        issue_idx = get_idx('ISSUE사항', '비고', '이슈', '특이사항', 'REMARK', 'NOTE', default=16)

        # 날짜 형식 최적화 헬퍼 함수
        def format_date_string(date_str):
            if pd.isna(date_str) or str(date_str).lower() == 'nan' or not str(date_str).strip():
                return ''
            s = str(date_str).replace('\n', ' ').strip()
            
            # 시간 부분 제거
            if ' 00:00:00' in s:
                s = s.split(' ')[0]
            elif 'T00:00:00' in s:
                s = s.split('T')[0]
                
            # YYYY-MM-DD 형식일 경우 연도 제거 (MM-DD만 남김)
            import re
            match = re.search(r'^\d{4}-(\d{2}-\d{2})$', s)
            if match:
                s = match.group(1)
            elif '/' in s: # YYYY/MM/DD 형식 처리
                match_slash = re.search(r'^\d{4}/(\d{2}/\d{2})$', s)
                if match_slash:
                    s = match_slash.group(1).replace('/', '-')
                    
            return s
            
        def get_val(row, col_index, default_val=''):
            if col_index is not None and col_index != -1 and col_index < len(row) and pd.notna(row.iloc[col_index]):
                v = str(row.iloc[col_index]).strip()
                if v.lower() == 'nan': return default_val
                if v.endswith('.0') and v[:-2].isdigit(): v = v[:-2]
                return v
            return default_val

        data = []
        for idx, row in df.iterrows():
            if header_idx != -1 and idx <= header_idx:
                continue
            
            # 빈 행 무시 조건 완화 (필수 식별자나 고유 의미가 있는 데이터가 존재하는지 확인)
            valid_cell_count = sum(1 for v in row if pd.notna(v) and str(v).strip() and str(v).lower() != 'nan')
            if valid_cell_count < 3: # 최소 3개 이상의 데이터가 있는 행만 취급
                continue

            serial = ''
            if serial_idx_mapped != -1 and serial_idx_mapped < len(row):
                v = str(row.iloc[serial_idx_mapped]).strip()
                if v.lower() not in ['nan', 'none', '-', '']:
                    serial = v
            
            # 호기를 찾지 못했을 경우, 전후 컬럼에서 숫자/영문 조합(예: MV1234) 형식 추론
            if not serial:
                import re
                for p_idx in range(len(row)):
                    val = str(row.iloc[p_idx]).strip()
                    # 영문자+숫자 조합이거나 4자리 이상 숫자이면 호기일 가능성
                    if val and val.lower() not in ['nan', 'none', '-', '']:
                        if re.search(r'[A-Za-z]+\d+', val) or (val.isdigit() and len(val) >= 4):
                            serial = val
                            break
                            
                # 그래도 못찾았다면 기본 인덱스들 시도
                if not serial:
                    potential_indices = [6, 5, 7, 8, 4]
                    for p_idx in potential_indices:
                        if p_idx < len(row):
                            val = str(row.iloc[p_idx]).strip()
                            if val and val.lower() not in ['nan', 'none', '-', '']:
                                serial = val
                                break

            # 식별 가능한 데이터(호기 혹은 모델, 출하처 등)가 없으면 스킵
            if not serial and len(str(get_val(row, model_idx))) < 2 and len(str(get_val(row, customer_idx))) < 2:
                continue

            # 타이틀이나 병합 셀의 잔여물이 데이터로 들어오는 것 방지
            first_val = str(row.iloc[0]).strip().upper() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            second_val = str(row.iloc[1]).strip().upper() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            if 'NO' in first_val or '생산월' in second_val or '기종' in second_val:
                continue

            item = {
                'no': get_val(row, no_idx),
                'month': get_val(row, month_idx),
                'section': get_val(row, section_idx),
                'model': get_val(row, model_idx).replace('\n', ' '),
                'serial': str(serial).replace('\n', ' '),
                'order': get_val(row, order_idx),
                'customer': get_val(row, customer_idx).replace('\n', ' '),
                'first_shipment': format_date_string(get_val(row, first_shipment_idx)),
                'target': format_date_string(get_val(row, target_idx)),
                'base': format_date_string(get_val(row, base_idx)),
                'first_start': format_date_string(get_val(row, first_start_idx)),
                'revised_start': format_date_string(get_val(row, revised_start_idx)),
                'nc': format_date_string(get_val(row, nc_idx)),
                'status': get_val(row, status_idx, '대기'),
                'issue': get_val(row, issue_idx).replace('\n', '<br>')
            }
            if not item['status']:
                item['status'] = '대기'
                
            data.append(item)
        
        print(f"Successfully parsed {len(data)} items.")
        return data
    except Exception as e:
        print(f"Error parsing excel: {e}")
        import traceback
        traceback.print_exc()
        return []

@app.errorhandler(404)
def page_not_found(e):
    # 사용자가 잘못된 주소(예: /1직 등)로 접속 시 404 에러 화면 대신 메인 화면으로 돌려보냅니다.
    return redirect(url_for('index'))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/work-report')
@app.route('/work_report')
def work_report():
    return render_template('work_report.html')

@app.route('/shop_calendar')
def shop_calendar():
    return render_template('shop_calendar.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        import datetime
        today_str = datetime.datetime.now().strftime("%Y.%m.%d")
        new_filename = f"업로드파일({today_str}).xlsx"
        save_path = os.path.join(BASE_DIR, new_filename)
        
        file.save(save_path)
        try:
            data = parse_production_data(save_path)
            return jsonify(data)
        except Exception as e:
            print(f"Error processing uploaded file: {e}")
            return jsonify({'error': str(e)}), 500

@app.route('/api/sheets')
def get_sheets():
    sheets = get_excel_sheets()
    return jsonify(sheets)

@app.route('/api/data')
def get_data():
    file_type = request.args.get('type', 'plan') # 'plan' or 'daily'
    sheet_name = request.args.get('sheet', None)
    
    if file_type == 'daily' and sheet_name:
        data = parse_production_data(EXCEL_FILE_DAILY, sheet_name=sheet_name)
    else:
        # 기본값은 출하 계획표 시트
        data = parse_production_data(EXCEL_FILE_PLAN, sheet_name=0)
    return jsonify(data)

if __name__ == '__main__':
    # 5000번 포트에서 실행
    print("공정 관리 시스템 서버를 시작합니다...")
    # 포트를 5003로 변경하여 기존 프로세스와의 충돌 방지 및 브라우저 캐시 완벽 우회
    app.run(debug=True, port=5003, threaded=True)
